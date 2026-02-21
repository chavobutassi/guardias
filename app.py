"""
APLICACIÓN WEB - GESTOR DE GUARDIAS 2026 (VERSIÓN INTEGRADA)
=============================================================
Backend Flask con generador de calendario integrado y sistema mejorado de gestión.

Mejoras v4.0:
- ✅ Generador de calendario integrado en la app
- ✅ Creación automática del archivo Excel al iniciar
- ✅ Sugerencias automáticas SOLO para personas activas
- ✅ Validación de fechas en asignaciones
- ✅ Indicadores visuales de disponibilidad
- ✅ Filtros inteligentes por estado
- ✅ Alertas de conflictos de disponibilidad

Autor: Sistema de Gestión de Guardias
Versión: 4.0 - Con Generador Integrado
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import json
from datetime import datetime, date, timedelta
from collections import defaultdict
import streamlit as st

app = Flask(__name__)
CORS(app)

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

EXCEL_FILE = "calendario_guardias_2026.xlsx"
HISTORIAL_FILE = "historial_guardias.json"
DISPONIBILIDAD_FILE = "disponibilidad.json"

# ============================================================================
# TABLA DE PERSONAL — fuente única de verdad
# orden_llenado : quién llena guardia primero (1 = más antiguo, llena antes)
# nombre        : clave interna usada en Excel, JSON, historial
# rina          : nro. de matrícula RINA (acredita antigüedad sin discusión)
# ============================================================================
PERSONAS_INFO = [
    {"orden": 1,  "nombre": "TNIM BUTASSI",        "rina": 1490},
    {"orden": 2,  "nombre": "TNAU BARRIOS",        "rina": 1512},
    {"orden": 3,  "nombre": "TN MACHUCA",          "rina": 1516},
    {"orden": 4,  "nombre": "TF ZALAZAR",          "rina": 1650},
    {"orden": 5,  "nombre": "TF ONETO CAJAL",      "rina": 1789},
    {"orden": 6,  "nombre": "TFCO LEDESMA",        "rina": 1840},
    {"orden": 7,  "nombre": "TFIM GONZALEZ",       "rina": 1855},
    {"orden": 8,  "nombre": "TFIM RACEDO BRITOS",  "rina": 2065},
    {"orden": 9,  "nombre": "TCCO PALMA",          "rina": 2093},
    {"orden": 10, "nombre": "TC LEDESMA",          "rina": 2142},
    {"orden": 11, "nombre": "GUIM DIAZ",           "rina": 2240},
    {"orden": 12, "nombre": "GUIM TORRES",         "rina": 2260},
    {"orden": 13, "nombre": "GUCO BENITEZ",        "rina": 2281},
]

# Lista de nombres limpios EN EL ORDEN DE LLENADO (más moderno primero).
# Esta lista es la clave interna usada en todo el sistema.
PERSONAS = [p["nombre"] for p in PERSONAS_INFO]

# Diccionarios de acceso rápido
PERSONA_ORDEN = {p["nombre"]: p["orden"] for p in PERSONAS_INFO}
PERSONA_RINA  = {p["nombre"]: p["rina"]  for p in PERSONAS_INFO}

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

MAP_MESES = {mes: idx + 1 for idx, mes in enumerate(MESES)}

# Feriados nacionales Argentina 2026
FERIADOS_2026 = {
    date(2026, 1, 1),   # Año Nuevo
    date(2026, 2, 16), date(2026, 2, 17),  # Carnaval
    date(2026, 3, 23), date(2026, 3, 24),  # Memoria, Verdad y Justicia
    date(2026, 4, 2),   # Malvinas
    date(2026, 4, 3),   # Viernes Santo
    date(2026, 5, 1),   # Día del Trabajador
    date(2026, 5, 25),  # Revolución de Mayo
    date(2026, 6, 15), date(2026, 6, 20),  # Paso a la Inmortalidad del Gral. Belgrano
    date(2026, 7, 9), date(2026, 7, 10),   # Independencia
    date(2026, 8, 17),  # Paso a la Inmortalidad del Gral. San Martín
    date(2026, 10, 12), # Día del Respeto a la Diversidad Cultural
    date(2026, 11, 23), # Día de la Soberanía Nacional
    date(2026, 12, 7), date(2026, 12, 8),  # Inmaculada Concepción
    date(2026, 12, 25)  # Navidad
}

# Convertir feriados a strings para búsquedas
FERIADOS_2026_STR = {f.strftime("%Y-%m-%d") for f in FERIADOS_2026}

# ============================================================================
# GENERADOR DE CALENDARIO INTEGRADO
# ============================================================================

def generar_calendario_guardias_2026():
    """
    Genera el calendario 2026 tipo grilla semanal con:
    - Feriados en rojo
    - Fin de semana en rojo
    - Lunes a jueves en azul
    - Viernes en amarillo
    - Vísperas de feriado en amarillo
    
    Estructura:
    - Fila 1: Encabezados (Lun, Mar, Mié, Jue, Vie, Sáb, Dom)
    - Filas pares (2,4,6...): Números de días
    - Filas impares (3,5,7...): Asignaciones de personas
    """
    print("📅 Generando calendario 2026...")
    
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    months = [
        ("Enero", 1), ("Febrero", 2), ("Marzo", 3), ("Abril", 4),
        ("Mayo", 5), ("Junio", 6), ("Julio", 7), ("Agosto", 8),
        ("Septiembre", 9), ("Octubre", 10), ("Noviembre", 11), ("Diciembre", 12)
    ]

    for mname, mnum in months:
        ws = wb.create_sheet(mname)
        
        # Fila 1: Encabezados
        ws.append(["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"])

        d = date(2026, mnum, 1)
        week = [None] * 7
        start_wd = d.weekday()

        # Llenar primera semana
        for i in range(start_wd, 7):
            week[i] = d
            d += timedelta(days=1)

        row_num = 2  # Empezar en fila 2
        
        while True:
            # Fila con números de día
            ws.append([c.day if c else "" for c in week])
            numero_row = ws[ws.max_row]
            
            # Fila para asignaciones (vacía inicialmente)
            ws.append([""] * 7)
            
            # Aplicar colores a las celdas de números
            for col_idx, cell in enumerate(numero_row, start=1):
                if not cell.value:
                    continue

                cell_date = date(2026, mnum, cell.value)

                # Feriado o fin de semana
                if cell_date in FERIADOS_2026 or cell_date.weekday() >= 5:
                    cell.fill = fill_red
                # Viernes o víspera
                elif cell_date.weekday() == 4 or (cell_date + timedelta(days=1)) in FERIADOS_2026:
                    cell.fill = fill_yellow
                else:
                    cell.fill = fill_blue

            # Preparar siguiente semana
            week = [None] * 7
            for i in range(7):
                if d.month == mnum:
                    week[i] = d
                    d += timedelta(days=1)

            if not any(week):
                break

    wb.save(EXCEL_FILE)
    print(f"✅ Calendario generado: {EXCEL_FILE}")



def inicializar_calendario():
    """
    Inicializa el calendario al arrancar la aplicación.
    Solo genera si no existe o si el usuario lo solicita.
    """
    if not os.path.exists(EXCEL_FILE):
        print("📋 No se encontró archivo de calendario. Generando...")
        generar_calendario_guardias_2026()
    else:
        print(f"✅ Calendario encontrado: {EXCEL_FILE}")


# ============================================================================
# FUNCIONES DE DISPONIBILIDAD
# ============================================================================

def cargar_disponibilidad():
    """Carga el estado de disponibilidad desde archivo JSON"""
    if not os.path.exists(DISPONIBILIDAD_FILE):
        # Crear archivo inicial con todos activos
        disponibilidad = {
            persona: {
                "activo": True,
                "motivo": None,
                "desde": None,
                "hasta": None
            } for persona in PERSONAS
        }
        guardar_disponibilidad(disponibilidad)
        return disponibilidad
    
    with open(DISPONIBILIDAD_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def guardar_disponibilidad(disponibilidad):
    """Guarda el estado de disponibilidad en archivo JSON"""
    with open(DISPONIBILIDAD_FILE, 'w', encoding='utf-8') as f:
        json.dump(disponibilidad, f, indent=2, ensure_ascii=False)


def persona_disponible(persona, fecha=None):
    """
    Verifica si una persona está disponible en una fecha específica.
    
    Args:
        persona: Nombre de la persona
        fecha: Fecha a verificar (string YYYY-MM-DD, objeto date, o None para verificación general)
    
    Returns:
        bool: True si está disponible, False si no
    """
    disponibilidad = cargar_disponibilidad()
    
    if persona not in disponibilidad:
        return True
    
    info = disponibilidad[persona]
    
    # Si está marcado como activo, está disponible
    if info['activo']:
        return True
    
    # Si está inactivo sin fechas, NO está disponible nunca
    if not info['desde'] and not info['hasta']:
        return False
    
    # Si no se proporciona fecha, retornar estado general
    if fecha is None:
        return False
    
    # Verificar si la fecha está en el rango de inactividad
    try:
        if isinstance(fecha, str):
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
        elif isinstance(fecha, date):
            fecha_obj = fecha
        else:
            return False
        
        # Verificar límite inferior (desde)
        if info['desde']:
            desde = datetime.strptime(info['desde'], "%Y-%m-%d").date()
            if fecha_obj < desde:
                return True  # Fecha anterior a inicio de inactividad
        
        # Verificar límite superior (hasta)
        if info['hasta']:
            hasta = datetime.strptime(info['hasta'], "%Y-%m-%d").date()
            if fecha_obj > hasta:
                return True  # Fecha posterior a fin de inactividad
        
        # Está dentro del rango de inactividad
        return False
        
    except Exception as e:
        print(f"Error al verificar disponibilidad: {e}")
        return not info['activo']


def obtener_personas_activas(fecha=None):
    """
    Retorna lista de personas activas en una fecha específica.
    
    Args:
        fecha: Fecha a verificar (opcional)
    
    Returns:
        list: Lista de personas disponibles
    """
    return [p for p in PERSONAS if persona_disponible(p, fecha)]


def get_motivo_indisponibilidad(persona, fecha=None):
    """
    Obtiene el motivo de indisponibilidad de una persona.
    
    Returns:
        str or None: Motivo si está indisponible, None si está disponible
    """
    disponibilidad = cargar_disponibilidad()
    
    if persona not in disponibilidad:
        return None
    
    if persona_disponible(persona, fecha):
        return None
    
    return disponibilidad[persona].get('motivo', 'No especificado')


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def tipo_dia_calendario(anio, mes, dia):
    """
    Determina el tipo de día: hábil, víspera o feriado
    """
    try:
        fecha = date(anio, mes, dia)
    except ValueError:
        return "habil"
    
    weekday = fecha.weekday()  # 0=lun ... 6=dom
    fecha_str = fecha.strftime("%Y-%m-%d")

    # Sábados y domingos
    if weekday in (5, 6):
        return "feriado"

    # Feriados oficiales
    if fecha_str in FERIADOS_2026_STR:
        return "feriado"

    # Víspera: viernes o día antes de feriado
    if weekday == 4:  # Viernes
        return "vispera"
    
    # Día antes de feriado (lun-jue)
    manana = fecha + timedelta(days=1)
    if manana.strftime("%Y-%m-%d") in FERIADOS_2026_STR and weekday <= 4:
        return "vispera"

    # Día hábil normal
    return "habil"


def obtener_dias_del_mes_mejorado(hoja, mes_nombre):
    """
    Detección robusta de días en formato calendario grid.
    Busca cualquier número entero que sea un día válido del mes.
    """
    # Mapeo de días de la semana a español
    dias_semana_map = {
        'Lun': 'Lunes', 'Mar': 'Martes', 'Mié': 'Miércoles', 'Miercoles': 'Miércoles',
        'Jue': 'Jueves', 'Vie': 'Viernes', 'Sáb': 'Sábado', 'Sab': 'Sábado', 
        'Dom': 'Domingo',
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }
    
    dias = {}
    anio = 2026
    mes_num = MAP_MESES.get(mes_nombre)

    if not mes_num:
        return {}

    # Buscar en toda la hoja (primeras 50 filas y 10 columnas deberían ser suficiente)
    dias_encontrados = {}  # {dia_num: (row, col)}
    
    for row in range(1, 50):
        for col in range(1, 10):
            celda = hoja.cell(row=row, column=col)
            
            # Detectar número de día
            if isinstance(celda.value, int) and 1 <= celda.value <= 31:
                dia_num = celda.value
                
                # Validar que sea un día real del mes
                try:
                    fecha_dia = date(anio, mes_num, dia_num)
                except ValueError:
                    continue
                
                # Guardar solo la primera ocurrencia de cada día
                if dia_num not in dias_encontrados:
                    dias_encontrados[dia_num] = (row, col)
    
    # Procesar cada día encontrado
    for dia_num, (row, col) in dias_encontrados.items():
        try:
            fecha_dia = date(anio, mes_num, dia_num)
        except ValueError:
            continue
        
        tipo_dia = tipo_dia_calendario(anio, mes_num, dia_num)
        
        # Buscar celda de asignación
        # Probar en la celda de abajo primero (formato de 2 filas)
        fila_asignacion = row + 1
        col_letra = get_column_letter(col)
        celda_ref = f"{col_letra}{fila_asignacion}"
        
        # Leer valor de asignación
        celda_asignacion = hoja.cell(row=fila_asignacion, column=col)
        valor = celda_asignacion.value
        
        # Verificar si el valor es una persona válida
        persona_actual = None
        if valor:
            valor_str = str(valor).strip()
            if valor_str in PERSONAS:
                persona_actual = valor_str
        
        # Determinar día de la semana
        # Intentar obtenerlo del encabezado
        dia_semana = fecha_dia.strftime("%A")
        encabezado = hoja.cell(row=1, column=col).value
        if encabezado:
            encabezado_str = str(encabezado).strip()
            if encabezado_str in dias_semana_map:
                dia_semana = dias_semana_map[encabezado_str]
        
        # Mapear día en inglés a español
        dias_en_es = {
            'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
            'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
        }
        dia_semana = dias_en_es.get(dia_semana, dia_semana)
        
        # VERIFICAR DISPONIBILIDAD
        fecha_str = fecha_dia.strftime("%Y-%m-%d")
        disponible = persona_disponible(persona_actual, fecha_str) if persona_actual else True
        motivo_indisponible = None if disponible else get_motivo_indisponibilidad(persona_actual, fecha_str)
        
        dias[dia_num] = {
            "tipo": tipo_dia,
            "celda_ref": celda_ref,
            "persona": persona_actual,
            "dia_semana": dia_semana,
            "fecha": fecha_str,
            "disponible": disponible,
            "motivo_indisponible": motivo_indisponible
        }
    
    return dias


# ============================================================================
# LÓGICA DE SUGERENCIAS
# ============================================================================

def sugerir_persona_para_dia_mejorado(wb, mes, dia_num, excluir=[]):
    """
    Sugiere la mejor persona para un día considerando SOLO personas activas.
    
    CAMBIO CLAVE: Solo considera personas que estén disponibles en esa fecha específica.
    """
    if mes not in MAP_MESES:
        return None
    
    mes_num = MAP_MESES[mes]
    anio = 2026
    
    try:
        fecha = date(anio, mes_num, dia_num)
        fecha_str = fecha.strftime("%Y-%m-%d")
    except ValueError:
        return None
    
    # OBTENER SOLO PERSONAS ACTIVAS EN ESA FECHA ESPECÍFICA
    personas_disponibles = obtener_personas_activas(fecha_str)
    
    # Excluir las personas que ya tienen guardia ese día o están en la lista de exclusión
    personas_disponibles = [p for p in personas_disponibles if p not in excluir]
    
    if not personas_disponibles:
        return None
    
    # Contar guardias hasta el mes anterior
    contador = defaultdict(int)
    
    for persona in personas_disponibles:
        contador[persona] = 0
    
    # Contar guardias en meses anteriores
    idx_mes_actual = MESES.index(mes)
    for mes_previo in MESES[:idx_mes_actual]:
        if mes_previo not in wb.sheetnames:
            continue
        
        hoja_previo = wb[mes_previo]
        dias_previos = obtener_dias_del_mes_mejorado(hoja_previo, mes_previo)
        
        for info in dias_previos.values():
            if info.get('persona') in contador:
                contador[info['persona']] += 1
    
    # Contar guardias en el mes actual (días anteriores)
    if mes in wb.sheetnames:
        hoja = wb[mes]
        dias_mes = obtener_dias_del_mes_mejorado(hoja, mes)
        
        for dia, info in dias_mes.items():
            if dia < dia_num and info.get('persona') in contador:
                contador[info['persona']] += 1
    
    # Retornar la persona con menos guardias
    if contador:
        persona_sugerida = min(contador.items(), key=lambda x: x[1])[0]
        return persona_sugerida
    
    return None


# ============================================================================
# CÁLCULO DE DISTRIBUCIÓN
# ============================================================================

def calcular_distribucion_planificada_mejorada(wb, solo_activos=True):
    """
    Calcula la distribución planificada de guardias por mes.
    """
    distribucion = {}
    acumulado_real = defaultdict(int)
    acumulado_ideal = defaultdict(float)
    
    # Determinar qué personas considerar
    if solo_activos:
        personas_considerar = obtener_personas_activas()
    else:
        personas_considerar = PERSONAS.copy()
    
    for mes in MESES:
        if mes not in wb.sheetnames:
            continue
        
        hoja = wb[mes]
        dias_mes = obtener_dias_del_mes_mejorado(hoja, mes)
        
        # Contar guardias asignadas en el mes
        guardias_mes = defaultdict(int)
        for info in dias_mes.values():
            persona = info.get('persona')
            if persona:
                guardias_mes[persona] += 1
                acumulado_real[persona] += 1
        
        # Calcular ideal del mes
        total_dias = len(dias_mes)
        
        # Contar solo personas activas día por día
        personas_activas_mes = set()
        for dia_num, info in dias_mes.items():
            fecha = info.get('fecha')
            activos_dia = obtener_personas_activas(fecha)
            personas_activas_mes.update(activos_dia)
        
        num_personas_activas = len(personas_activas_mes)
        
        if num_personas_activas > 0:
            ideal_mes = total_dias / num_personas_activas
        else:
            ideal_mes = 0
        
        # Actualizar acumulado ideal
        for persona in personas_activas_mes:
            acumulado_ideal[persona] += ideal_mes
        
        # Preparar datos del mes
        distribucion[mes] = {
            "total_dias": total_dias,
            "personas_activas": num_personas_activas,
            "ideal_mes": round(ideal_mes, 2),
            "distribucion": {}
        }
        
        # Datos por persona
        for persona in personas_considerar:
            real_mes = guardias_mes.get(persona, 0)
            real_acum = acumulado_real.get(persona, 0)
            ideal_acum = acumulado_ideal.get(persona, 0)
            diferencia = real_acum - ideal_acum
            
            # Determinar estado
            if abs(diferencia) < 0.5:
                estado = "equilibrado"
            elif diferencia > 0:
                estado = "hizo_mas"
            else:
                estado = "debe_mas"
            
            distribucion[mes]["distribucion"][persona] = {
                "real_mes": real_mes,
                "ideal_mes": round(ideal_mes, 2),
                "acumulado_real": real_acum,
                "acumulado_ideal": round(ideal_acum, 2),
                "diferencia_acumulada": round(diferencia, 2),
                "estado": estado,
                "activo": persona_disponible(persona)
            }
    
    return distribucion


# ============================================================================
# FUNCIÓN DE HISTORIAL
# ============================================================================

def registrar_en_historial(evento):
    """Registra un evento en el historial"""
    try:
        if os.path.exists(HISTORIAL_FILE):
            with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
                historial = json.load(f)
        else:
            historial = []
        
        evento['timestamp'] = datetime.now().isoformat()
        historial.append(evento)
        
        with open(HISTORIAL_FILE, 'w', encoding='utf-8') as f:
            json.dump(historial, f, indent=2, ensure_ascii=False)
    except:
        pass


# ============================================================================
# ENDPOINTS API
# ============================================================================

@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')


@app.route('/api/generar-calendario', methods=['POST'])
def regenerar_calendario():
    """Regenera el calendario (útil si se corrompe o se quiere resetear)"""
    try:
        generar_calendario_guardias_2026()
        return jsonify({
            "success": True,
            "mensaje": "✅ Calendario regenerado exitosamente",
            "archivo": EXCEL_FILE
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/disponibilidad', methods=['GET'])
def get_disponibilidad():
    """Obtiene el estado completo de disponibilidad"""
    try:
        disponibilidad = cargar_disponibilidad()
        
        # Enriquecer con información de estado
        for persona in disponibilidad:
            disponibilidad[persona]['disponible_hoy'] = persona_disponible(persona)
            disponibilidad[persona]['orden'] = PERSONA_ORDEN.get(persona, 99)
            disponibilidad[persona]['rina'] = PERSONA_RINA.get(persona)
        
        return jsonify(disponibilidad)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/disponibilidad/<persona>', methods=['PUT'])
def update_disponibilidad(persona):
    """Actualiza la disponibilidad de una persona"""
    try:
        if persona not in PERSONAS:
            return jsonify({"error": "Persona no encontrada"}), 404
        
        data = request.json
        disponibilidad = cargar_disponibilidad()
        
        # Actualizar campos
        disponibilidad[persona]['activo'] = data.get('activo', disponibilidad[persona]['activo'])
        disponibilidad[persona]['motivo'] = data.get('motivo')
        disponibilidad[persona]['desde'] = data.get('desde')
        disponibilidad[persona]['hasta'] = data.get('hasta')
        
        guardar_disponibilidad(disponibilidad)
        
        # Registrar en historial
        registrar_en_historial({
            "accion": "cambio_disponibilidad",
            "persona": persona,
            "activo": disponibilidad[persona]['activo'],
            "motivo": disponibilidad[persona]['motivo'],
            "desde": disponibilidad[persona]['desde'],
            "hasta": disponibilidad[persona]['hasta']
        })
        
        return jsonify({
            "success": True,
            "mensaje": f"✓ Disponibilidad actualizada para {persona}",
            "disponibilidad": disponibilidad[persona]
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/personas/activas')
def get_personas_activas():
    """Lista personas activas (con opción de filtrar por fecha)"""
    try:
        fecha = request.args.get('fecha')  # Formato: YYYY-MM-DD
        activas = obtener_personas_activas(fecha)
        
        disponibilidad = cargar_disponibilidad()
        resultado = []
        
        for persona in activas:
            info = disponibilidad.get(persona, {})
            resultado.append({
                "nombre": persona,
                "activo": info.get('activo', True),
                "motivo": info.get('motivo'),
                "desde": info.get('desde'),
                "hasta": info.get('hasta'),
                "orden": PERSONA_ORDEN.get(persona, 99),
                "rina": PERSONA_RINA.get(persona)
            })
        
        return jsonify({
            "fecha": fecha or "general",
            "total": len(resultado),
            "personas": resultado
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/validar-asignacion', methods=['POST'])
def validar_asignacion():
    """Valida si una persona puede ser asignada a una fecha específica"""
    try:
        data = request.json
        persona = data.get('persona')
        mes = data.get('mes')
        dia = data.get('dia')
        
        if not all([persona, mes, dia]):
            return jsonify({"error": "Faltan parámetros"}), 400
        
        if persona not in PERSONAS:
            return jsonify({"error": "Persona no encontrada"}), 404
        
        if mes not in MAP_MESES:
            return jsonify({"error": "Mes no válido"}), 400
        
        # Construir fecha
        mes_num = MAP_MESES[mes]
        anio = 2026
        
        try:
            fecha = date(anio, mes_num, dia)
            fecha_str = fecha.strftime("%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Fecha no válida"}), 400
        
        # Verificar disponibilidad
        disponible = persona_disponible(persona, fecha_str)
        
        resultado = {
            "valido": disponible,
            "persona": persona,
            "fecha": fecha_str,
            "mes": mes,
            "dia": dia
        }
        
        if not disponible:
            resultado["motivo_rechazo"] = get_motivo_indisponibilidad(persona, fecha_str)
            resultado["advertencia"] = f"⚠️ {persona} no está disponible el {fecha_str}"
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/calendario')
def get_calendario():
    """Endpoint mejorado que incluye información de disponibilidad"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        meses_disponibles = [hoja for hoja in wb.sheetnames if hoja in MESES]
        wb.close()
        
        # Incluir información de personas activas
        personas_activas = obtener_personas_activas()
        
        return jsonify({
            "meses": meses_disponibles,
            "personas": PERSONAS,
            "personas_activas": personas_activas,
            "total_personas": len(PERSONAS),
            "total_activas": len(personas_activas),
            "total_inactivas": len(PERSONAS) - len(personas_activas)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/mes/<mes>')
def get_mes(mes):
    """Endpoint mejorado con validación de disponibilidad"""
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        print(f"\n📅 Cargando mes: {mes}")
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            print(f"❌ Mes '{mes}' no encontrado en hojas: {wb.sheetnames}")
            return jsonify({"error": f"Mes '{mes}' no encontrado en el archivo"}), 404
        
        hoja = wb[mes]
        print(f"✓ Hoja '{mes}' encontrada")
        
        # Debug: mostrar primeras celdas
        print(f"  Fila 1: {[hoja.cell(1, c).value for c in range(1, 8)]}")
        print(f"  Fila 2: {[hoja.cell(2, c).value for c in range(1, 8)]}")
        
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        wb.close()
        
        print(f"✓ Días detectados: {len(dias)}")
        if len(dias) > 0:
            print(f"  Ejemplo - Día 1: {dias.get(1, 'No encontrado')}")
        
        # Contar estadísticas
        total_dias = len(dias)
        asignados = sum(1 for d in dias.values() if d.get('persona'))
        pendientes = total_dias - asignados
        
        # Contar conflictos de disponibilidad
        conflictos = sum(1 for d in dias.values() if d.get('persona') and not d.get('disponible'))
        
        # Contar por tipo
        tipos = {"habil": 0, "vispera": 0, "feriado": 0}
        for dia_info in dias.values():
            tipos[dia_info['tipo']] += 1
        
        return jsonify({
            "mes": mes,
            "dias": dias,
            "estadisticas": {
                "total": total_dias,
                "asignados": asignados,
                "pendientes": pendientes,
                "conflictos": conflictos,
                "habiles": tipos.get("habil", 0),
                "visperas": tipos.get("vispera", 0),
                "feriados": tipos.get("feriado", 0)
            }
        })
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"\n❌ ERROR en get_mes({mes}):")
        print(error_detail)
        return jsonify({
            "error": str(e),
            "detail": error_detail
        }), 500


@app.route('/api/asignar', methods=['POST'])
def asignar_guardia():
    """Asignar guardia con validación de disponibilidad"""
    try:
        data = request.json
        mes = data.get('mes')
        dia = data.get('dia')
        persona = data.get('persona')
        forzar = data.get('forzar', False)  # Permitir forzar asignación
        
        if not all([mes, dia, persona]):
            return jsonify({"error": "Faltan parámetros"}), 400
        
        if mes not in MESES or persona not in PERSONAS:
            return jsonify({"error": "Mes o persona no válidos"}), 400
        
        # Construir fecha y validar disponibilidad
        mes_num = MAP_MESES[mes]
        try:
            fecha = date(2026, mes_num, dia)
            fecha_str = fecha.strftime("%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Día no válido"}), 400
        
        # Verificar disponibilidad
        if not forzar:
            if not persona_disponible(persona, fecha_str):
                motivo = get_motivo_indisponibilidad(persona, fecha_str)
                return jsonify({
                    "error": "persona_no_disponible",
                    "mensaje": f"⚠️ {persona} no está disponible el {fecha_str}",
                    "motivo": motivo,
                    "sugerencia": "Use 'forzar: true' para asignar de todas formas"
                }), 400
        
        # Continuar con asignación normal
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        if dia not in dias:
            wb.close()
            return jsonify({"error": "Día no encontrado"}), 404
        
        celda_ref = dias[dia]['celda_ref']
        persona_anterior = dias[dia].get('persona')
        
        # Escribir en Excel
        hoja[celda_ref] = persona
        wb.save(EXCEL_FILE)
        wb.close()
        
        # Registrar en historial
        registrar_en_historial({
            "accion": "asignar",
            "mes": mes,
            "dia": dia,
            "antes": persona_anterior,
            "despues": persona,
            "forzado": forzar and not persona_disponible(persona, fecha_str)
        })
        
        mensaje = f"✓ Guardia asignada: {persona} el día {dia} de {mes}"
        if forzar:
            mensaje += " (⚠️ FORZADO - persona no disponible)"
        
        return jsonify({
            "success": True,
            "mensaje": mensaje,
            "dia": dia,
            "persona": persona,
            "anterior": persona_anterior,
            "forzado": forzar
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/sugerir/<mes>/<int:dia>')
def sugerir_persona(mes, dia):
    """Sugerir persona SOLO entre activos"""
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        # Construir fecha para verificar disponibilidad
        mes_num = MAP_MESES[mes]
        try:
            fecha = date(2026, mes_num, dia)
            fecha_str = fecha.strftime("%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Día no válido"}), 400
        
        wb = load_workbook(EXCEL_FILE)
        
        # Usar función mejorada que solo considera activos
        sugerencia = sugerir_persona_para_dia_mejorado(wb, mes, dia)
        
        wb.close()
        
        if sugerencia:
            # Verificar disponibilidad (doble check)
            disponible = persona_disponible(sugerencia, fecha_str)
            
            return jsonify({
                "sugerencia": sugerencia,
                "disponible": disponible,
                "fecha": fecha_str,
                "advertencia": None if disponible else f"⚠️ Persona sugerida no está disponible"
            })
        else:
            return jsonify({
                "sugerencia": None,
                "mensaje": "No hay personas disponibles para este día",
                "personas_activas": obtener_personas_activas(fecha_str)
            })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/distribucion/planificada')
def distribucion_planificada():
    """Distribución planificada con opción de incluir inactivos"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        solo_activos = request.args.get('solo_activos', 'true').lower() == 'true'
        mes_especifico = request.args.get('mes')
        
        wb = load_workbook(EXCEL_FILE)
        distribucion = calcular_distribucion_planificada_mejorada(wb, solo_activos=solo_activos)
        wb.close()
        
        if mes_especifico:
            if mes_especifico not in MESES:
                return jsonify({"error": f"Mes '{mes_especifico}' no válido"}), 400
            
            if mes_especifico in distribucion:
                return jsonify({
                    "mes": mes_especifico,
                    "solo_activos": solo_activos,
                    **distribucion[mes_especifico]
                })
            else:
                return jsonify({"error": f"Mes '{mes_especifico}' no encontrado"}), 404
        
        return jsonify({
            "solo_activos": solo_activos,
            "distribucion": distribucion
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/info')
def get_info():
    """Endpoint de información general del sistema"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo Excel no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        meses_disponibles = [hoja for hoja in wb.sheetnames if hoja in MESES]
        wb.close()
        
        # Obtener información de disponibilidad
        activos = obtener_personas_activas()
        
        return jsonify({
            "status": "ok",
            "version": "4.0",
            "meses": meses_disponibles,
            "personas": PERSONAS,
            "personas_activas": activos,
            "total_personas": len(PERSONAS),
            "total_activas": len(activos),
            "total_inactivas": len(PERSONAS) - len(activos),
            "excel_file": EXCEL_FILE,
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/health')
def health_check():
    """Health check mejorado"""
    activos = obtener_personas_activas()
    
    return jsonify({
        "status": "ok",
        "version": "4.0 - Con Generador Integrado",
        "timestamp": datetime.now().isoformat(),
        "excel_exists": os.path.exists(EXCEL_FILE),
        "disponibilidad_exists": os.path.exists(DISPONIBILIDAD_FILE),
        "personas_total": len(PERSONAS),
        "personas_activas": len(activos),
        "personas_inactivas": len(PERSONAS) - len(activos),
        "mejoras": [
            "✅ Generador de calendario integrado",
            "✅ Creación automática al iniciar",
            "✅ Sugerencias solo para personas activas",
            "✅ Validación de disponibilidad en asignaciones",
            "✅ Detección de conflictos de disponibilidad"
        ]
    })


@app.route('/api/historial')
def get_historial():
    """Obtiene el historial de cambios"""
    try:
        if not os.path.exists(HISTORIAL_FILE):
            return jsonify({"historial": []})
        
        with open(HISTORIAL_FILE, 'r', encoding='utf-8') as f:
            historial = json.load(f)
        
        # Retornar últimos 100 registros
        return jsonify({
            "historial": historial[-100:],
            "total": len(historial)
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/eliminar', methods=['POST'])
def eliminar_guardia():
    """Elimina una guardia asignada"""
    try:
        data = request.json
        mes = data.get('mes')
        dia = data.get('dia')
        
        if not all([mes, dia]):
            return jsonify({"error": "Faltan parámetros"}), 400
        
        # Convertir día a int si viene como string
        try:
            dia = int(dia)
        except (ValueError, TypeError):
            return jsonify({"error": "Día debe ser un número"}), 400
        
        if mes not in MESES:
            return jsonify({"error": "Mes no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        if dia not in dias:
            wb.close()
            return jsonify({"error": f"Día {dia} no encontrado en {mes}"}), 404
        
        celda_ref = dias[dia]['celda_ref']
        persona_anterior = dias[dia].get('persona')
        
        if not persona_anterior:
            wb.close()
            return jsonify({"error": "No hay guardia asignada para eliminar"}), 400
        
        # Eliminar (vaciar celda)
        hoja[celda_ref] = None
        wb.save(EXCEL_FILE)
        wb.close()
        
        # Registrar en historial
        registrar_en_historial({
            "accion": "eliminar",
            "mes": mes,
            "dia": dia,
            "persona": persona_anterior
        })
        
        return jsonify({
            "success": True,
            "mensaje": f"✓ Guardia eliminada: {persona_anterior} del día {dia} de {mes}",
            "dia": dia,
            "persona_eliminada": persona_anterior
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/distribucion/auto/<mes>', methods=['POST'])
def distribucion_automatica(mes):
    """
    Distribución automática COMPLETA Y EQUITATIVA de guardias para un mes.
    
    GARANTIZA:
    1. TODO el mes queda cubierto (100% de días asignados)
    2. Distribución equitativa considerando "peso" de cada tipo de día
    3. Balance entre hábiles, vísperas y feriados para cada persona
    
    SISTEMA DE PUNTOS:
    - Día hábil = 1 punto
    - Víspera = 1.5 puntos (más pesado)
    - Feriado = 2 puntos (más pesado)
    
    Cada persona recibe aproximadamente los mismos puntos totales.
    """
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        # Obtener personas activas para este mes
        mes_num = MAP_MESES[mes]
        personas_activas_mes = set()
        for dia_num, info in dias.items():
            fecha = info.get('fecha')
            activos_dia = obtener_personas_activas(fecha)
            personas_activas_mes.update(activos_dia)
        
        personas_lista = sorted(list(personas_activas_mes), key=lambda p: PERSONA_ORDEN.get(p, 99))
        num_personas = len(personas_lista)
        
        if num_personas == 0:
            wb.close()
            return jsonify({"error": "No hay personas activas disponibles"}), 400
        
        # Separar días por tipo
        dias_por_tipo = {
            'habil': [],
            'vispera': [],
            'feriado': []
        }
        
        for dia_num, info in dias.items():
            dias_por_tipo[info['tipo']].append(dia_num)
        
        # Ordenar días
        for tipo in dias_por_tipo:
            dias_por_tipo[tipo].sort()
        
        # Sistema de puntos por tipo de día
        PUNTOS = {
            'habil': 1.0,
            'vispera': 1.5,
            'feriado': 2.0
        }
        
        # Calcular puntos totales del mes
        puntos_totales = (
            len(dias_por_tipo['habil']) * PUNTOS['habil'] +
            len(dias_por_tipo['vispera']) * PUNTOS['vispera'] +
            len(dias_por_tipo['feriado']) * PUNTOS['feriado']
        )
        
        # Puntos ideales por persona
        puntos_por_persona = puntos_totales / num_personas
        
        # Inicializar tracking
        asignaciones = {}
        puntos_acumulados = {p: 0.0 for p in personas_lista}
        dias_asignados = {p: {'habil': 0, 'vispera': 0, 'feriado': 0, 'total': 0} for p in personas_lista}
        
        # Crear lista de todos los días con su tipo y puntos
        todos_los_dias = []
        for tipo in ['feriado', 'vispera', 'habil']:  # Orden: más pesados primero
            for dia_num in dias_por_tipo[tipo]:
                todos_los_dias.append({
                    'dia': dia_num,
                    'tipo': tipo,
                    'puntos': PUNTOS[tipo],
                    'fecha': dias[dia_num]['fecha']
                })
        
        # ALGORITMO DE DISTRIBUCIÓN EQUITATIVA
        # Asignar cada día a la persona que menos puntos acumulados tenga
        for dia_info in todos_los_dias:
            dia_num = dia_info['dia']
            tipo = dia_info['tipo']
            puntos = dia_info['puntos']
            fecha = dia_info['fecha']
            
            # Encontrar persona disponible con menos puntos
            personas_disponibles = [
                p for p in personas_lista 
                if persona_disponible(p, fecha)
            ]
            
            if not personas_disponibles:
                # Ninguna persona disponible, buscar la menos ocupada de todas
                personas_disponibles = personas_lista
            
            # Ordenar por puntos acumulados (menos puntos primero)
            personas_disponibles.sort(key=lambda p: puntos_acumulados[p])
            
            # Asignar a la primera (la que tiene menos puntos)
            persona_elegida = personas_disponibles[0]
            
            asignaciones[dia_num] = persona_elegida
            puntos_acumulados[persona_elegida] += puntos
            dias_asignados[persona_elegida][tipo] += 1
            dias_asignados[persona_elegida]['total'] += 1
        
        # Aplicar asignaciones al Excel
        # Primero, limpiar todas las asignaciones del mes
        for dia_num, info in dias.items():
            celda_ref = info['celda_ref']
            hoja[celda_ref] = None
        
        # Luego, aplicar nuevas asignaciones
        cambios = 0
        for dia_num, persona in asignaciones.items():
            celda_ref = dias[dia_num]['celda_ref']
            hoja[celda_ref] = persona
            cambios += 1
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        # Registrar en historial
        registrar_en_historial({
            "accion": "distribucion_automatica_completa",
            "mes": mes,
            "cambios": cambios,
            "puntos_sistema": PUNTOS,
            "personas": personas_lista
        })
        
        # Calcular estadísticas finales
        conteo = {}
        for persona in personas_lista:
            conteo[persona] = {
                'habil': dias_asignados[persona]['habil'],
                'vispera': dias_asignados[persona]['vispera'],
                'feriado': dias_asignados[persona]['feriado'],
                'total': dias_asignados[persona]['total'],
                'puntos': round(puntos_acumulados[persona], 2)
            }
        
        # Calcular desviación (qué tan equitativo quedó)
        puntos_values = list(puntos_acumulados.values())
        puntos_min = min(puntos_values)
        puntos_max = max(puntos_values)
        diferencia_max = puntos_max - puntos_min
        
        return jsonify({
            "success": True,
            "mensaje": f"✅ Distribución completa y equitativa para {mes}",
            "mes": mes,
            "dias_asignados": cambios,
            "dias_totales": len(dias),
            "cobertura": "100%",
            "personas_participantes": num_personas,
            "sistema_puntos": {
                "habil": f"{PUNTOS['habil']} punto",
                "vispera": f"{PUNTOS['vispera']} puntos",
                "feriado": f"{PUNTOS['feriado']} puntos"
            },
            "puntos_totales_mes": round(puntos_totales, 2),
            "puntos_ideal_por_persona": round(puntos_por_persona, 2),
            "equidad": {
                "puntos_minimos": round(puntos_min, 2),
                "puntos_maximos": round(puntos_max, 2),
                "diferencia": round(diferencia_max, 2),
                "nivel": "excelente" if diferencia_max < 2 else "bueno" if diferencia_max < 4 else "aceptable"
            },
            "distribucion": conteo
        })
        
    except Exception as e:
        import traceback
        print(f"Error en distribución automática: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/asignar/usuario/<mes>/<int:dia>', methods=['POST'])
def asignar_usuario_a_dia(mes, dia):
    """
    Permite que un usuario se auto-asigne a un día específico.
    Verifica que el día esté disponible y que no exceda su guardia.
    """
    try:
        data = request.json
        persona = data.get('persona')
        
        if not persona:
            return jsonify({"error": "Falta el nombre de la persona"}), 400
        
        if persona not in PERSONAS:
            return jsonify({"error": "Persona no encontrada"}), 404
        
        if mes not in MESES:
            return jsonify({"error": "Mes no válido"}), 400
        
        # Construir fecha y validar disponibilidad
        mes_num = MAP_MESES[mes]
        try:
            fecha = date(2026, mes_num, dia)
            fecha_str = fecha.strftime("%Y-%m-%d")
        except ValueError:
            return jsonify({"error": "Día no válido"}), 400
        
        # Verificar disponibilidad
        if not persona_disponible(persona, fecha_str):
            motivo = get_motivo_indisponibilidad(persona, fecha_str)
            return jsonify({
                "error": "no_disponible",
                "mensaje": f"No estás disponible el {fecha_str}",
                "motivo": motivo
            }), 400
        
        # Verificar que el día no esté ya asignado
        wb = load_workbook(EXCEL_FILE)
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        if dia not in dias:
            wb.close()
            return jsonify({"error": "Día no encontrado"}), 404
        
        if dias[dia].get('persona'):
            persona_actual = dias[dia]['persona']
            wb.close()
            return jsonify({
                "error": "dia_ocupado",
                "mensaje": f"Este día ya está asignado a {persona_actual}"
            }), 400
        
        # Asignar
        celda_ref = dias[dia]['celda_ref']
        hoja[celda_ref] = persona
        wb.save(EXCEL_FILE)
        wb.close()
        
        # Registrar en historial
        registrar_en_historial({
            "accion": "auto_asignacion",
            "mes": mes,
            "dia": dia,
            "persona": persona,
            "fecha": fecha_str
        })
        
        return jsonify({
            "success": True,
            "mensaje": f"✅ Te asignaste exitosamente al día {dia} de {mes}",
            "dia": dia,
            "persona": persona,
            "tipo_dia": dias[dia]['tipo']
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/estadisticas/usuario/<persona>')
def estadisticas_usuario(persona):
    """
    Obtiene las estadísticas de asignaciones de un usuario específico.
    Muestra cuántos días hábiles, vísperas y feriados tiene asignados.
    """
    try:
        if persona not in PERSONAS:
            return jsonify({"error": "Persona no encontrada"}), 404
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        stats = {
            "total": 0,
            "habil": 0,
            "vispera": 0,
            "feriado": 0,
            "por_mes": {}
        }
        
        for mes in MESES:
            if mes not in wb.sheetnames:
                continue
            
            hoja = wb[mes]
            dias = obtener_dias_del_mes_mejorado(hoja, mes)
            
            mes_stats = {
                "total": 0,
                "habil": 0,
                "vispera": 0,
                "feriado": 0,
                "dias": []
            }
            
            for dia_num, info in dias.items():
                if info.get('persona') == persona:
                    tipo = info['tipo']
                    stats['total'] += 1
                    stats[tipo] += 1
                    mes_stats['total'] += 1
                    mes_stats[tipo] += 1
                    mes_stats['dias'].append({
                        "dia": dia_num,
                        "tipo": tipo,
                        "fecha": info['fecha'],
                        "dia_semana": info['dia_semana']
                    })
            
            if mes_stats['total'] > 0:
                stats['por_mes'][mes] = mes_stats
        
        wb.close()
        
        return jsonify({
            "persona": persona,
            "estadisticas": stats,
            "activo": persona_disponible(persona)
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/distribucion/balancear/<mes>', methods=['POST'])
def distribucion_balancear(mes):
    """
    Distribución inteligente que SOLO llena días pendientes.
    Prioriza a personas con MENOS guardias para lograr equidad.
    
    Parámetros:
        solo_calcular (bool): Si es True, solo calcula y muestra sin aplicar cambios
    
    Perfecto para casos como Febrero donde ya hay días asignados (1-17)
    y solo se quieren llenar los pendientes (18-28) de forma equitativa.
    """
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        # Verificar si solo queremos calcular sin aplicar
        data = request.json or {}
        solo_calcular = data.get('solo_calcular', False)
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        # Obtener personas activas
        mes_num = MAP_MESES[mes]
        personas_activas_mes = set()
        for dia_num, info in dias.items():
            fecha = info.get('fecha')
            activos_dia = obtener_personas_activas(fecha)
            personas_activas_mes.update(activos_dia)
        
        personas_lista = sorted(list(personas_activas_mes), key=lambda p: PERSONA_ORDEN.get(p, 99))
        num_personas = len(personas_lista)
        
        if num_personas == 0:
            wb.close()
            return jsonify({"error": "No hay personas activas disponibles"}), 400
        
        # Contar guardias ya asignadas por persona
        PUNTOS = {'habil': 1.0, 'vispera': 1.5, 'feriado': 2.0}
        
        conteo_actual = {p: {'total': 0, 'puntos': 0.0, 'habil': 0, 'vispera': 0, 'feriado': 0} for p in personas_lista}
        dias_pendientes = []
        
        for dia_num, info in dias.items():
            persona = info.get('persona')
            tipo = info['tipo']
            
            if persona and persona in conteo_actual:
                # Día ya asignado
                conteo_actual[persona]['total'] += 1
                conteo_actual[persona][tipo] += 1
                conteo_actual[persona]['puntos'] += PUNTOS[tipo]
            else:
                # Día pendiente
                dias_pendientes.append({
                    'dia': dia_num,
                    'tipo': tipo,
                    'puntos': PUNTOS[tipo],
                    'fecha': info['fecha']
                })
        
        # Ordenar días pendientes por puntos (más pesados primero)
        dias_pendientes.sort(key=lambda x: x['puntos'], reverse=True)
        
        print(f"\n📊 DISTRIBUCIÓN BALANCEADA - {mes} (Solo calcular: {solo_calcular})")
        print(f"Días pendientes: {len(dias_pendientes)}")
        print(f"Personas activas: {num_personas}")
        
        # Estado inicial
        print("\n📋 Estado Inicial:")
        for persona in sorted(conteo_actual.keys(), key=lambda p: conteo_actual[p]['total']):
            datos = conteo_actual[persona]
            print(f"  {persona}: {datos['total']} días ({datos['puntos']:.1f} pts)")
        
        # Asignar cada día pendiente a quien tenga MENOS
        asignaciones_nuevas = {}
        
        for dia_info in dias_pendientes:
            dia_num = dia_info['dia']
            tipo = dia_info['tipo']
            puntos = dia_info['puntos']
            fecha = dia_info['fecha']
            
            # Encontrar personas disponibles
            personas_disponibles = [
                p for p in personas_lista 
                if persona_disponible(p, fecha)
            ]
            
            if not personas_disponibles:
                personas_disponibles = personas_lista
            
            # Ordenar por puntos acumulados (menos primero)
            personas_disponibles.sort(key=lambda p: conteo_actual[p]['puntos'])
            
            # Asignar al primero (quien tiene menos)
            persona_elegida = personas_disponibles[0]
            
            asignaciones_nuevas[dia_num] = {
                'persona': persona_elegida,
                'tipo': tipo,
                'puntos': puntos
            }
            
            # Actualizar conteo
            conteo_actual[persona_elegida]['total'] += 1
            conteo_actual[persona_elegida][tipo] += 1
            conteo_actual[persona_elegida]['puntos'] += puntos
        
        # SOLO APLICAR SI NO ES "solo_calcular"
        cambios = 0
        if not solo_calcular:
            for dia_num, asig in asignaciones_nuevas.items():
                celda_ref = dias[dia_num]['celda_ref']
                hoja[celda_ref] = asig['persona']
                cambios += 1
            
            wb.save(EXCEL_FILE)
            print("\n💾 Cambios APLICADOS al Excel")
        else:
            cambios = len(asignaciones_nuevas)
            print("\n📋 Cambios CALCULADOS (no aplicados)")
        
        wb.close()
        
        # Estado final
        print("\n📋 Estado Final (proyectado):")
        for persona in sorted(conteo_actual.keys(), key=lambda p: conteo_actual[p]['total'], reverse=True):
            datos = conteo_actual[persona]
            print(f"  {persona}: {datos['total']} días ({datos['puntos']:.1f} pts) - H:{datos['habil']} V:{datos['vispera']} F:{datos['feriado']}")
        
        # Calcular diferencia
        puntos_values = [conteo_actual[p]['puntos'] for p in personas_lista]
        puntos_min = min(puntos_values)
        puntos_max = max(puntos_values)
        diferencia = puntos_max - puntos_min
        
        # Registrar en historial solo si se aplicó
        if not solo_calcular:
            registrar_en_historial({
                "accion": "distribucion_balanceada",
                "mes": mes,
                "cambios": cambios,
                "diferencia_final": round(diferencia, 2)
            })
        
        mensaje_base = "calculada" if solo_calcular else "completada"
        
        return jsonify({
            "success": True,
            "mensaje": f"✅ Distribución balanceada {mensaje_base} para {mes}",
            "mes": mes,
            "solo_calcular": solo_calcular,
            "dias_pendientes_asignados": cambios,
            "dias_que_ya_estaban": len(dias) - len(dias_pendientes),
            "personas_participantes": num_personas,
            "estado_final": {
                persona: {
                    'total': conteo_actual[persona]['total'],
                    'habil': conteo_actual[persona]['habil'],
                    'vispera': conteo_actual[persona]['vispera'],
                    'feriado': conteo_actual[persona]['feriado'],
                    'puntos': round(conteo_actual[persona]['puntos'], 1)
                }
                for persona in personas_lista
            },
            "equidad": {
                "puntos_minimos": round(puntos_min, 1),
                "puntos_maximos": round(puntos_max, 1),
                "diferencia": round(diferencia, 1),
                "nivel": "excelente" if diferencia < 2 else "bueno" if diferencia < 4 else "aceptable"
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error en distribución balanceada: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/reporte/anual')
def reporte_anual():
    """
    Genera reporte anual completo con estadísticas de todos los meses.
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        # Estadísticas generales
        dias_totales = 0
        dias_asignados = 0
        
        # Por persona
        por_persona = {}
        
        # Por mes
        meses_data = {}
        
        for mes in MESES:
            if mes not in wb.sheetnames:
                continue
            
            hoja = wb[mes]
            dias = obtener_dias_del_mes_mejorado(hoja, mes)
            
            total_mes = len(dias)
            asignados_mes = sum(1 for d in dias.values() if d.get('persona'))
            
            dias_totales += total_mes
            dias_asignados += asignados_mes
            
            meses_data[mes] = {
                'total': total_mes,
                'asignados': asignados_mes,
                'pendientes': total_mes - asignados_mes
            }
            
            # Contar por persona
            for dia_info in dias.values():
                persona = dia_info.get('persona')
                if persona:
                    if persona not in por_persona:
                        por_persona[persona] = {
                            'total': 0,
                            'por_mes': {}
                        }
                    
                    por_persona[persona]['total'] += 1
                    
                    if mes not in por_persona[persona]['por_mes']:
                        por_persona[persona]['por_mes'][mes] = 0
                    
                    por_persona[persona]['por_mes'][mes] += 1
        
        wb.close()
        
        # Calcular porcentaje de cobertura
        porcentaje_cobertura = round((dias_asignados / dias_totales * 100), 1) if dias_totales > 0 else 0
        
        return jsonify({
            'totales': {
                'dias_totales': dias_totales,
                'dias_asignados': dias_asignados,
                'dias_pendientes': dias_totales - dias_asignados,
                'porcentaje_cobertura': f"{porcentaje_cobertura}"
            },
            'meses': meses_data,
            'por_persona': por_persona
        })
        
    except Exception as e:
        import traceback
        print(f"Error generando reporte anual: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/mes/<mes>/resetear', methods=['POST'])
def resetear_mes(mes):
    """
    Resetea/limpia todas las asignaciones de un mes específico.
    Útil para empezar de cero sin afectar otros meses.
    """
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        # Contar cuántas asignaciones hay
        asignaciones_eliminadas = 0
        personas_afectadas = set()
        
        # Limpiar todas las celdas de asignación
        for dia_num, info in dias.items():
            persona = info.get('persona')
            if persona:
                celda_ref = info['celda_ref']
                hoja[celda_ref] = None
                asignaciones_eliminadas += 1
                personas_afectadas.add(persona)
        
        wb.save(EXCEL_FILE)
        wb.close()
        
        # Registrar en historial
        registrar_en_historial({
            "accion": "resetear_mes",
            "mes": mes,
            "asignaciones_eliminadas": asignaciones_eliminadas,
            "personas_afectadas": list(personas_afectadas)
        })
        
        return jsonify({
            "success": True,
            "mensaje": f"✅ Mes {mes} reseteado completamente",
            "mes": mes,
            "asignaciones_eliminadas": asignaciones_eliminadas,
            "personas_afectadas": len(personas_afectadas),
            "detalle": f"Se eliminaron {asignaciones_eliminadas} guardias de {len(personas_afectadas)} personas"
        })
        
    except Exception as e:
        import traceback
        print(f"Error al resetear mes: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/guardias/sugeridas/<mes>', methods=['GET'])
def calcular_guardias_sugeridas(mes):
    """
    Calcula guardias sugeridas para cada persona SIN asignar automáticamente.
    Muestra cuántos días de cada tipo debería tener cada persona para balance equitativo.
    """
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        wb.close()
        
        # Obtener personas activas
        mes_num = MAP_MESES[mes]
        personas_activas_mes = set()
        for dia_num, info in dias.items():
            fecha = info.get('fecha')
            activos_dia = obtener_personas_activas(fecha)
            personas_activas_mes.update(activos_dia)
        
        personas_lista = sorted(list(personas_activas_mes), key=lambda p: PERSONA_ORDEN.get(p, 99))
        num_personas = len(personas_lista)
        
        if num_personas == 0:
            return jsonify({"error": "No hay personas activas disponibles"}), 400
        
        # Contar días ya asignados por persona
        asignados_actuales = {p: {'habil': 0, 'vispera': 0, 'feriado': 0, 'total': 0, 'puntos': 0.0} for p in personas_lista}
        dias_disponibles_por_tipo = {
            'habil': [],
            'vispera': [],
            'feriado': []
        }
        
        PUNTOS = {'habil': 1.0, 'vispera': 1.5, 'feriado': 2.0}
        
        for dia_num, info in dias.items():
            tipo = info['tipo']
            persona = info.get('persona')
            
            if persona and persona in asignados_actuales:
                asignados_actuales[persona][tipo] += 1
                asignados_actuales[persona]['total'] += 1
                asignados_actuales[persona]['puntos'] += PUNTOS[tipo]
            else:
                # Día disponible
                dias_disponibles_por_tipo[tipo].append(dia_num)
        
        # Calcular guardia ideal total
        total_dias = len(dias)
        guardia_ideal = total_dias / num_personas
        
        # Calcular puntos totales del mes
        puntos_totales = sum(len(dias_disponibles_por_tipo[t]) * PUNTOS[t] for t in ['habil', 'vispera', 'feriado'])
        for persona_data in asignados_actuales.values():
            puntos_totales += persona_data['puntos']
        
        puntos_ideal_por_persona = puntos_totales / num_personas
        
        # Calcular cuántos días más necesita cada persona
        guardias_sugeridas = {}
        for persona in personas_lista:
            actual = asignados_actuales[persona]
            faltan_dias = max(0, guardia_ideal - actual['total'])
            faltan_puntos = max(0, puntos_ideal_por_persona - actual['puntos'])
            
            guardias_sugeridas[persona] = {
                'actuales': actual,
                'guardia_ideal_total': round(guardia_ideal, 1),
                'faltan_dias': round(faltan_dias, 1),
                'puntos_ideal': round(puntos_ideal_por_persona, 1),
                'faltan_puntos': round(faltan_puntos, 1),
                'balance': 'OK' if abs(faltan_dias) < 1 else 'NECESITA_MAS' if faltan_dias > 0 else 'TIENE_DEMAS'
            }
        
        return jsonify({
            "success": True,
            "mes": mes,
            "total_dias_mes": total_dias,
            "dias_asignados": total_dias - sum(len(v) for v in dias_disponibles_por_tipo.values()),
            "dias_disponibles": sum(len(v) for v in dias_disponibles_por_tipo.values()),
            "disponibles_por_tipo": {
                "habil": len(dias_disponibles_por_tipo['habil']),
                "vispera": len(dias_disponibles_por_tipo['vispera']),
                "feriado": len(dias_disponibles_por_tipo['feriado'])
            },
            "guardia_ideal_por_persona": round(guardia_ideal, 1),
            "puntos_ideal_por_persona": round(puntos_ideal_por_persona, 1),
            "personas_activas": num_personas,
            "guardias_sugeridas": guardias_sugeridas
        })
        
    except Exception as e:
        import traceback
        print(f"Error en guardias sugeridas: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/cuotas/sugeridas/<mes>', methods=['GET'])
def calcular_cuotas_sugeridas(mes):
    """
    Calcula las cuotas sugeridas EXACTAS para cada persona usando el algoritmo de distribución automática.
    SOLO SIMULA - NO guarda en Excel.
    """
    try:
        if mes not in MESES:
            return jsonify({"error": f"Mes '{mes}' no válido"}), 400
        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        wb = load_workbook(EXCEL_FILE)
        
        if mes not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Mes '{mes}' no encontrado"}), 404
        
        hoja = wb[mes]
        dias = obtener_dias_del_mes_mejorado(hoja, mes)
        
        # NO cerrar wb todavía - lo necesitamos para simular
        
        # Obtener personas activas
        mes_num = MAP_MESES[mes]
        personas_activas_mes = set()
        for dia_num, info in dias.items():
            fecha = info.get('fecha')
            activos_dia = obtener_personas_activas(fecha)
            personas_activas_mes.update(activos_dia)
        
        personas_lista = sorted(list(personas_activas_mes), key=lambda p: PERSONA_ORDEN.get(p, 99))
        num_personas = len(personas_lista)
        
        if num_personas == 0:
            wb.close()
            return jsonify({"error": "No hay personas activas disponibles"}), 400
        
        # ============================================================================
        # USAR EL MISMO ALGORITMO DE DISTRIBUCIÓN AUTOMÁTICA (MODO SIMULACIÓN)
        # ============================================================================
        
        PUNTOS = {'habil': 1.0, 'vispera': 1.5, 'feriado': 2.0}
        
        # Contar asignaciones ACTUALES (antes de simular)
        asignados_antes = {p: {'habil': 0, 'vispera': 0, 'feriado': 0, 'total': 0, 'puntos': 0.0} for p in personas_lista}
        
        for dia_num, info in dias.items():
            persona = info.get('persona')
            tipo = info['tipo']
            
            if persona and persona in asignados_antes:
                asignados_antes[persona][tipo] += 1
                asignados_antes[persona]['total'] += 1
                asignados_antes[persona]['puntos'] += PUNTOS[tipo]
        
        # SIMULAR distribución completa (como distribución_automatica pero sin guardar)
        # Crear copia de hoja para simular
        from copy import deepcopy
        dias_simulados = deepcopy(dias)
        
        # Resetear todas las asignaciones en la simulación
        for dia_num in dias_simulados:
            dias_simulados[dia_num]['persona'] = None
        
        # Algoritmo de distribución (igual que distribucion_automatica)
        conteo_simulado = {p: {'total': 0, 'puntos': 0.0, 'habil': 0, 'vispera': 0, 'feriado': 0} for p in personas_lista}
        
        # Ordenar días por peso (feriado > víspera > hábil)
        dias_ordenados = sorted(
            dias_simulados.items(),
            key=lambda x: PUNTOS[x[1]['tipo']],
            reverse=True
        )
        
        # Inicializar conteo para DNRD
        conteo_simulado['DNRD'] = {'total': 0, 'puntos': 0.0, 'habil': 0, 'vispera': 0, 'feriado': 0, 'dias': []}
        
        # Asignar cada día a quien tenga menos puntos
        for dia_num, info in dias_ordenados:
            tipo = info['tipo']
            fecha = info['fecha']
            puntos = PUNTOS[tipo]
            
            # Encontrar personas disponibles
            personas_disponibles = [
                p for p in personas_lista
                if persona_disponible(p, fecha)
            ]
            
            if not personas_disponibles:
                # Nadie disponible -> DNRD cubre este día
                conteo_simulado['DNRD']['total'] += 1
                conteo_simulado['DNRD'][tipo] += 1
                conteo_simulado['DNRD']['puntos'] += puntos
                conteo_simulado['DNRD']['dias'].append({'dia': dia_num, 'fecha': fecha, 'tipo': tipo})
                continue
            
            # Ordenar por puntos acumulados (menos primero)
            personas_disponibles.sort(key=lambda p: (conteo_simulado[p]['puntos'], conteo_simulado[p]['total']))
            
            # Asignar al primero
            persona_elegida = personas_disponibles[0]
            
            # Actualizar conteo simulado
            conteo_simulado[persona_elegida]['total'] += 1
            conteo_simulado[persona_elegida][tipo] += 1
            conteo_simulado[persona_elegida]['puntos'] += puntos
        
        wb.close()
        
        # ============================================================================
        # CALCULAR SUGERENCIAS (diferencia entre simulado y actual)
        # ============================================================================
        
        total_dias = len(dias)
        cuota_ideal = total_dias / num_personas
        puntos_totales = sum(conteo_simulado[p]['puntos'] for p in personas_lista)
        puntos_ideal = puntos_totales / num_personas
        
        cuotas_sugeridas = {}
        for persona in personas_lista:
            antes = asignados_antes[persona]
            despues = conteo_simulado[persona]
            
            # Calcular sugeridos (diferencia)
            sugerido_habil = despues['habil'] - antes['habil']
            sugerido_vispera = despues['vispera'] - antes['vispera']
            sugerido_feriado = despues['feriado'] - antes['feriado']
            
            # Determinar balance
            diferencia = despues['total'] - cuota_ideal
            
            if abs(diferencia) < 0.5:
                balance = 'OK'
            elif diferencia < 0:
                balance = 'NECESITA_MAS'
            else:
                balance = 'TIENE_DEMAS'
            
            cuotas_sugeridas[persona] = {
                'actuales': {
                    'habil': antes['habil'],
                    'vispera': antes['vispera'],
                    'feriado': antes['feriado'],
                    'total': antes['total'],
                    'puntos': round(antes['puntos'], 1)
                },
                'sugeridos': {
                    'habil': sugerido_habil,
                    'vispera': sugerido_vispera,
                    'feriado': sugerido_feriado,
                    'total': sugerido_habil + sugerido_vispera + sugerido_feriado
                },
                'proyectado': {
                    'habil': despues['habil'],
                    'vispera': despues['vispera'],
                    'feriado': despues['feriado'],
                    'total': despues['total'],
                    'puntos': round(despues['puntos'], 1)
                },
                'cuota_ideal_total': round(cuota_ideal, 1),
                'puntos_ideal': round(puntos_ideal, 1),
                'balance': balance,
                'orden': PERSONA_ORDEN.get(persona, 99),
                'rina': PERSONA_RINA.get(persona)
            }
        
        # Agregar info de DNRD si hubo días sin personal disponible
        dnrd_info = conteo_simulado.get('DNRD', {'total': 0, 'habil': 0, 'vispera': 0, 'feriado': 0, 'puntos': 0.0, 'dias': []})
        if dnrd_info['total'] > 0:
            cuotas_sugeridas['DNRD'] = {
                'actuales': {'habil': 0, 'vispera': 0, 'feriado': 0, 'total': 0, 'puntos': 0.0},
                'sugeridos': {
                    'habil': dnrd_info['habil'],
                    'vispera': dnrd_info['vispera'],
                    'feriado': dnrd_info['feriado'],
                    'total': dnrd_info['total']
                },
                'proyectado': {
                    'habil': dnrd_info['habil'],
                    'vispera': dnrd_info['vispera'],
                    'feriado': dnrd_info['feriado'],
                    'total': dnrd_info['total'],
                    'puntos': round(dnrd_info['puntos'], 1)
                },
                'dias_especificos': dnrd_info.get('dias', []),
                'cuota_ideal_total': 0,
                'puntos_ideal': 0,
                'balance': 'DNRD',
                'es_dnrd': True
            }

        # Logs para verificación
        print(f"\n📊 CUOTAS SUGERIDAS - {mes}")
        print(f"Total días: {total_dias}")
        print(f"Personas activas: {num_personas}")
        print(f"Cuota ideal: {cuota_ideal:.1f} días/persona")
        if dnrd_info['total'] > 0:
            print(f"  ⚠️  DNRD: {dnrd_info['total']} días sin personal disponible")
        
        total_proyectado = sum(c['proyectado']['total'] for c in cuotas_sugeridas.values())
        print(f"\n✅ Total proyectado: {total_proyectado} (debe ser {total_dias})")
        
        for persona in sorted(personas_lista):
            c = cuotas_sugeridas[persona]
            print(f"  {persona}: {c['actuales']['total']} + {c['sugeridos']['total']} = {c['proyectado']['total']}")
        
        # Calcular disponibles
        dias_ya_asignados = sum(antes['total'] for antes in asignados_antes.values())
        dias_disponibles = total_dias - dias_ya_asignados
        
        return jsonify({
            "success": True,
            "mes": mes,
            "total_dias_mes": total_dias,
            "dias_asignados": dias_ya_asignados,
            "dias_disponibles": dias_disponibles,
            "disponibles_por_tipo": {
                "habil": sum(1 for d in dias.values() if d['tipo'] == 'habil' and not d.get('persona')),
                "vispera": sum(1 for d in dias.values() if d['tipo'] == 'vispera' and not d.get('persona')),
                "feriado": sum(1 for d in dias.values() if d['tipo'] == 'feriado' and not d.get('persona'))
            },
            "cuota_ideal_por_persona": round(cuota_ideal, 1),
            "puntos_ideal_por_persona": round(puntos_ideal, 1),
            "personas_activas": num_personas,
            "cuotas_sugeridas": cuotas_sugeridas
        })
        
    except Exception as e:
        import traceback
        print(f"Error en cuotas sugeridas: {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500
@app.route('/api/descargar')
def descargar_excel():
    """Descarga el archivo Excel actualizado"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Archivo no encontrado"}), 404
        
        return send_file(
            EXCEL_FILE,
            as_attachment=True,
            download_name=f"calendario_guardias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# INICIAR APLICACIÓN
# ============================================================================

if __name__ == '__main__':
    print("""
    ╔══════════════════════════════════════════════════════════╗
    ║                                                          ║
    ║     🎯 GESTOR WEB DE GUARDIAS 2026 - V4.0 🚀            ║
    ║                                                          ║
    ║  ✅ Generador de calendario integrado                   ║
    ║  ✅ Creación automática del Excel                       ║
    ║  ✅ Gestión inteligente de disponibilidad               ║
    ║  ✅ Sugerencias SOLO para personas activas              ║
    ║  ✅ Validación automática de fechas                     ║
    ║  ✅ Detección de conflictos                             ║
    ║  ✅ Filtros avanzados activo/inactivo                   ║
    ║                                                          ║
    ╚══════════════════════════════════════════════════════════╝
    """)
    
    # Inicializar calendario
    inicializar_calendario()
    
    # Obtener puerto desde variable de entorno (Render lo asigna automáticamente)
    port = int(os.environ.get('PORT', 5000))
    
    print(f"""
    🌐 Accede a la aplicación en:
       http://localhost:{port}
    
    📱 Desde otro dispositivo en la misma red:
       http://[TU_IP]:{port}
    
    ⏹️  Para detener: Ctrl+C
    """)
    
    # En producción (Render): debug=False
    # En desarrollo local: debug=True
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)

